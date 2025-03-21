import docx
import openpyxl
from openpyxl.utils.cell import get_column_letter


""" Сервис для экспорта данных """
class ExportService:
    def export_to_docx(self, filename, data):
        doc = docx.Document()
        for key, value in data.items():
            doc.add_paragraph(f"{key}: {value}")

        doc.save(filename)

    def export_to_xlsx(self, filename, data):
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            # Найти первую свободную строку
            row_index = sheet.max_row + 1
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            row_index = 1

        # Записываем заголовки в первую строку, если это новый файл
        if row_index == 1:
            col_index = 1
            for key in data.keys():
                sheet.cell(row=row_index, column=col_index, value=key)
                col_index += 1
            row_index += 1  # Переходим к строке для записи данных

        # Записываем значения во вторую строку
        col_index = 1
        for value in data.values():
            sheet.cell(row=row_index, column=col_index, value=value)
            col_index += 1

        wb.save(filename)